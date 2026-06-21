# -*- coding: utf-8 -*-
"""Reformat the two SauceDemo workbooks to the studi-kasus PDF field specs.
   - Test Cases : ID Test | Nama Skenario | Langkah Pengujian | Expected Result
                  | Hasil Aktual | Status | Catatan Pengujian   (PDF C.1 + C.2)
   - Bug Report : Bug ID | Severity | Priority | Deskripsi Defect
                  | Langkah Reproduksi | Expected Result | Actual Result  (PDF C.3)
                  ('Raised By' removed)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "132B4D"; WHITE = "FFFFFF"; ALT = "E9EDF4"
PASS_C = "1E8E3E"; FAIL_C = "D93A3A"; MED_C = "E08A1E"; LOW_C = "5A6372"; HIGH_C = "D93A3A"

thin = Side(style="thin", color="C9D2DF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
ALT_FILL = PatternFill("solid", fgColor=ALT)
HEAD_FONT = Font(bold=True, color=WHITE, name="Calibri", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def c(x):
    return "" if x is None else str(x).strip()


def norm_status(v):
    u = v.strip().lower()
    if u.startswith("pass"): return "PASS"
    if u.startswith("fail"): return "FAIL"
    if "block" in u: return "BLOCKED"
    if "retest" in u: return "RETEST"
    if "not" in u: return "NOT EXECUTED"
    return v.upper()


def norm_level(v):
    u = v.strip().lower()
    if u.startswith("h"): return "High"
    if u.startswith("m"): return "Medium"   # handles 'medum' typo
    if u.startswith("l"): return "Low"
    if u.startswith("c"): return "Critical"
    return v.title()


def style_sheet(ws, widths, status_col=None, level_cols=()):
    ws.freeze_panes = "A2"
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    for ri, row in enumerate(ws.iter_rows(), start=1):
        ws.row_dimensions[ri].height = 30 if ri == 1 else None
        for cell in row:
            cell.border = BORDER
            if ri == 1:
                cell.fill = HEAD_FILL; cell.font = HEAD_FONT; cell.alignment = CENTER
            else:
                cell.alignment = WRAP if cell.column not in (status_col,) + tuple(level_cols) else CENTER
                if ri % 2 == 1:
                    cell.fill = ALT_FILL
    # color status / level cells
    for ri in range(2, ws.max_row + 1):
        if status_col:
            cell = ws.cell(row=ri, column=status_col)
            val = c(cell.value).upper()
            color = {"PASS": PASS_C, "FAIL": FAIL_C, "BLOCKED": MED_C,
                     "RETEST": MED_C, "NOT EXECUTED": LOW_C}.get(val)
            if color:
                cell.font = Font(bold=True, color=color, name="Calibri", size=11)
        for lc in level_cols:
            cell = ws.cell(row=ri, column=lc)
            val = c(cell.value)
            color = {"High": HIGH_C, "Critical": HIGH_C, "Medium": MED_C, "Low": LOW_C}.get(val)
            if color:
                cell.font = Font(bold=True, color=color, name="Calibri", size=11)


# ============================================================ TEST CASES
src = openpyxl.load_workbook("SauceDemo (Test Cases).xlsx", data_only=True)
out = openpyxl.Workbook(); out.remove(out.active)
TC_HEAD = ["ID Test", "Nama Skenario", "Langkah Pengujian", "Expected Result",
           "Hasil Aktual", "Status", "Catatan Pengujian"]
for ws in src.worksheets:
    rows = list(ws.iter_rows(values_only=True))
    hdr = [c(h) for h in rows[0]]
    def idx(name):
        for i, h in enumerate(hdr):
            if name.lower() in h.lower():
                return i
        return None
    iId, iSc, iStep, iExp, iAct, iPf = (idx("TC ID"), idx("Test case"), idx("Steps"),
                                        idx("Expected"), idx("Actual"), idx("Pass"))
    nws = out.create_sheet(ws.title)
    nws.append(TC_HEAD)
    for r in rows[1:]:
        v = [c(x) for x in r]
        if iId is None or iId >= len(v) or not v[iId]:
            continue
        st = norm_status(v[iPf]) if iPf is not None and iPf < len(v) else ""
        act = v[iAct] if iAct is not None and iAct < len(v) else ""
        catatan = "Sesuai dengan expected result" if st == "PASS" else \
                  ("Hasil tidak sesuai — lihat Defect Report" if st == "FAIL" else "")
        if st == "FAIL" and not act:
            act = "Tidak sesuai expected result"
        nws.append([
            v[iId],
            v[iSc] if iSc is not None and iSc < len(v) else "",
            v[iStep] if iStep is not None and iStep < len(v) else "",
            v[iExp] if iExp is not None and iExp < len(v) else "",
            act, st, catatan,
        ])
    style_sheet(nws, [13, 30, 42, 38, 38, 11, 28], status_col=6)
out.save("SauceDemo (Test Cases).xlsx")
print("Test Cases reformatted:", out.sheetnames)

# ============================================================ BUG REPORT
src = openpyxl.load_workbook("SauceDemo (Bug Report).xlsx", data_only=True)
ws = src.active
rows = list(ws.iter_rows(values_only=True))
hdr = [c(h) for h in rows[0]]
def bidx(name):
    for i, h in enumerate(hdr):
        if name.lower() in h.lower():
            return i
    return None
iId, iSev, iPri, iTitle, iSteps, iExp, iAct = (bidx("Bug ID"), bidx("Severity"),
    bidx("Priority"), bidx("Bug Title"), bidx("Steps"), bidx("Expected"), bidx("Actual"))
out = openpyxl.Workbook(); nws = out.active; nws.title = "Defect Report"
BUG_HEAD = ["Bug ID", "Severity", "Priority", "Deskripsi Defect",
            "Langkah Reproduksi", "Expected Result", "Actual Result"]
nws.append(BUG_HEAD)
for r in rows[1:]:
    v = [c(x) for x in r]
    if iId is None or iId >= len(v) or not v[iId]:
        continue
    bid = v[iId]
    try:
        bid = str(int(float(bid)))   # 1.0 -> "1"
    except ValueError:
        pass
    nws.append([
        bid,
        norm_level(v[iSev]) if iSev is not None and iSev < len(v) else "",
        norm_level(v[iPri]) if iPri is not None and iPri < len(v) else "",
        v[iTitle] if iTitle is not None and iTitle < len(v) else "",
        v[iSteps] if iSteps is not None and iSteps < len(v) else "",
        v[iExp] if iExp is not None and iExp < len(v) else "",
        v[iAct] if iAct is not None and iAct < len(v) else "",
    ])
style_sheet(nws, [9, 12, 12, 40, 46, 34, 34], level_cols=(2, 3))
out.save("SauceDemo (Bug Report).xlsx")
print("Bug Report reformatted: rows =", nws.max_row - 1)
